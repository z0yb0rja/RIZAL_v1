from __future__ import annotations

import io
import json
import os
import uuid
from pathlib import Path
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.security import get_current_user_with_roles, has_any_role
from app.database import get_db
from app.models.department import Department
from app.models.import_job import BulkImportJob
from app.models.program import Program
from app.models.school import SchoolAuditLog
from app.models.user import User
from app.repositories.import_repository import ImportRepository
from app.schemas.import_job import (
    ImportErrorItem,
    ImportJobCreateResponse,
    ImportJobStatusResponse,
    ImportPreviewResponse,
    ImportPreviewRow,
    RetryFailedRowsRequest,
)
from app.services.import_validation_service import (
    EXPECTED_HEADERS,
    HeaderValidationError,
    ValidationContext,
    suggest_fixes,
    validate_and_transform_row,
    validate_headers,
)
from app.worker.celery_app import celery_app

router = APIRouter(prefix="/api/admin", tags=["admin-import"])


def get_current_admin_or_school_it(
    current_user: User = Depends(get_current_user_with_roles),
) -> User:
    if not has_any_role(current_user, ["admin", "school_IT"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin or School IT privileges required",
        )
    return current_user


def _append_import_audit_log(
    db: Session,
    *,
    current_user: User,
    status_value: str,
    details: dict,
    action: str = "student_bulk_import_attempt",
) -> None:
    school_id = getattr(current_user, "school_id", None)
    if school_id is None:
        return

    db.add(
        SchoolAuditLog(
            school_id=school_id,
            actor_user_id=current_user.id,
            action=action,
            status=status_value,
            details=json.dumps(details, default=str),
        )
    )


def _ensure_user_school(current_user: User) -> int:
    school_id = getattr(current_user, "school_id", None)
    if school_id is None:
        raise HTTPException(status_code=403, detail="User is not assigned to a school")
    return school_id


def _validate_upload_basics(
    *,
    file: UploadFile,
    current_user: User,
    db: Session,
    settings,
) -> tuple[str, int]:
    filename = (file.filename or "").strip()
    if not filename:
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="failed",
            details={"reason": "missing file name"},
        )
        db.commit()
        raise HTTPException(status_code=400, detail="File name is required")
    if not filename.lower().endswith(".xlsx"):
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="failed",
            details={"reason": "unsupported file extension", "filename": filename},
        )
        db.commit()
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    file.file.seek(0, os.SEEK_END)
    size_bytes = file.file.tell()
    file.file.seek(0)
    if size_bytes <= 0:
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="failed",
            details={"reason": "empty upload", "filename": filename},
        )
        db.commit()
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    max_size_bytes = settings.import_max_file_size_mb * 1024 * 1024
    if size_bytes > max_size_bytes:
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="failed",
            details={
                "reason": "file size exceeded",
                "filename": filename,
                "size_bytes": size_bytes,
                "max_size_bytes": max_size_bytes,
            },
        )
        db.commit()
        raise HTTPException(
            status_code=413,
            detail=f"File size exceeds limit of {settings.import_max_file_size_mb} MB",
        )
    return filename, size_bytes


def _queue_import_job_from_file_bytes(
    *,
    db: Session,
    settings,
    current_user: User,
    filename: str,
    file_bytes: bytes,
    size_bytes: int,
    retried_from_job_id: str | None = None,
) -> ImportJobCreateResponse:
    repo = ImportRepository(db)
    recent_job_count = repo.count_recent_jobs(
        created_by_user_id=current_user.id,
        window_seconds=settings.import_rate_limit_window_seconds,
    )
    if recent_job_count >= settings.import_rate_limit_count:
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="rate_limited",
            details={
                "reason": "rate limit exceeded",
                "filename": filename,
                "window_seconds": settings.import_rate_limit_window_seconds,
            },
        )
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many import requests. Please wait before uploading again.",
        )

    job_id = str(uuid.uuid4())
    target_school_id = _ensure_user_school(current_user)
    storage_dir = Path(settings.import_storage_dir) / "uploads"
    storage_dir.mkdir(parents=True, exist_ok=True)
    stored_file_path = storage_dir / f"{job_id}.xlsx"
    stored_file_path.write_bytes(file_bytes)

    job = BulkImportJob(
        id=job_id,
        created_by_user_id=current_user.id,
        target_school_id=target_school_id,
        status="pending",
        original_filename=filename,
        stored_file_path=str(stored_file_path),
    )

    repo.create_job(job)
    _append_import_audit_log(
        db,
        current_user=current_user,
        status_value="queued",
        details={
            "job_id": job_id,
            "filename": filename,
            "size_bytes": size_bytes,
            "retried_from_job_id": retried_from_job_id,
        },
        action="student_bulk_import_retry" if retried_from_job_id else "student_bulk_import_attempt",
    )
    db.commit()
    celery_app.send_task("app.worker.tasks.process_student_import_job", args=[job_id])

    return ImportJobCreateResponse(
        job_id=job_id,
        status="pending",
        retried_from_job_id=retried_from_job_id,
    )


def _build_validation_context(db: Session, target_school_id: int) -> ValidationContext:
    department_lookup = {
        name.strip().lower(): department_id
        for department_id, name in db.query(Department.id, Department.name).all()
    }
    course_lookup = {
        name.strip().lower(): program_id
        for program_id, name in db.query(Program.id, Program.name).all()
    }
    return ValidationContext(
        target_school_id=target_school_id,
        department_lookup=department_lookup,
        course_lookup=course_lookup,
    )


@router.get("/import-students/template")
def download_import_students_template(
    current_user: User = Depends(get_current_admin_or_school_it),
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    sheet.append(EXPECTED_HEADERS)
    sheet.append(
        [
            "STU-00001",
            "student1@example.edu",
            "Doe",
            "Jane",
            "A",
            "Computer Science",
            "BS Computer Science",
        ]
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="student_import_template.xlsx"'},
    )


@router.post("/import-students/preview", response_model=ImportPreviewResponse)
def preview_import_students(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    settings = get_settings()
    filename, _ = _validate_upload_basics(
        file=file,
        current_user=current_user,
        db=db,
        settings=settings,
    )
    target_school_id = _ensure_user_school(current_user)
    file.file.seek(0)
    file_bytes = file.file.read()

    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError) as exc:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {exc}") from exc

    try:
        sheet = workbook.active
        total_rows = max((sheet.max_row or 1) - 1, 0)
        context = _build_validation_context(db, target_school_id)

        row_iter = sheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            return ImportPreviewResponse(
                filename=filename,
                total_rows=0,
                valid_rows=0,
                invalid_rows=0,
                can_commit=False,
                rows=[
                    ImportPreviewRow(
                        row=1,
                        status="failed",
                        errors=["Missing header row"],
                        suggestions=["Download and use the latest template before importing."],
                        row_data=None,
                    )
                ],
            )

        try:
            validate_headers(header_row)
        except HeaderValidationError as exc:
            message = str(exc)
            return ImportPreviewResponse(
                filename=filename,
                total_rows=total_rows,
                valid_rows=0,
                invalid_rows=max(total_rows, 1),
                can_commit=False,
                rows=[
                    ImportPreviewRow(
                        row=1,
                        status="failed",
                        errors=[message],
                        suggestions=suggest_fixes([message]),
                        row_data=None,
                    )
                ],
            )

        valid_rows = 0
        invalid_rows = 0
        preview_rows: list[ImportPreviewRow] = []

        for row_number, row_values in enumerate(row_iter, start=2):
            transformed, row_errors, row_data = validate_and_transform_row(
                row_number=row_number,
                row_values=row_values,
                context=context,
            )
            if row_errors:
                invalid_rows += 1
                status_value = "failed"
                errors = row_errors
                suggestions = suggest_fixes(row_errors)
            else:
                valid_rows += 1
                status_value = "valid"
                errors = []
                suggestions = []

            if len(preview_rows) < 200:
                preview_rows.append(
                    ImportPreviewRow(
                        row=row_number,
                        status=status_value,
                        errors=errors,
                        suggestions=suggestions,
                        row_data=row_data if transformed is None else row_data,
                    )
                )

        return ImportPreviewResponse(
            filename=filename,
            total_rows=total_rows,
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            can_commit=(total_rows > 0 and invalid_rows == 0),
            rows=preview_rows,
        )
    finally:
        workbook.close()


@router.post("/import-students", response_model=ImportJobCreateResponse)
def import_students(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    settings = get_settings()
    filename, size_bytes = _validate_upload_basics(
        file=file,
        current_user=current_user,
        db=db,
        settings=settings,
    )
    file.file.seek(0)
    file_bytes = file.file.read()
    return _queue_import_job_from_file_bytes(
        db=db,
        settings=settings,
        current_user=current_user,
        filename=filename,
        file_bytes=file_bytes,
        size_bytes=size_bytes,
    )


@router.post("/import-students/retry-failed/{job_id}", response_model=ImportJobCreateResponse)
def retry_failed_rows(
    job_id: str,
    payload: RetryFailedRowsRequest | None = None,
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    repo = ImportRepository(db)
    parent_job = repo.get_job(job_id)
    if not parent_job or parent_job.created_by_user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Import job not found")

    errors = repo.fetch_errors(job_id, limit=10000)
    if not errors:
        raise HTTPException(status_code=400, detail="No failed rows available to retry")

    selected_rows = set((payload.row_numbers if payload else []) or [])
    retry_row_payloads: list[dict] = []
    for item in errors:
        if selected_rows and item.row_number not in selected_rows:
            continue
        if isinstance(item.row_data, dict):
            retry_row_payloads.append(item.row_data)

    if not retry_row_payloads:
        raise HTTPException(
            status_code=400,
            detail="No retryable row payloads found for the selected rows",
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students-Retry"
    sheet.append(EXPECTED_HEADERS)
    for row_data in retry_row_payloads:
        sheet.append([str(row_data.get(header, "")) for header in EXPECTED_HEADERS])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    file_bytes = output.read()

    retry_filename = f"retry_{parent_job.original_filename or 'student_import.xlsx'}"
    settings = get_settings()
    return _queue_import_job_from_file_bytes(
        db=db,
        settings=settings,
        current_user=current_user,
        filename=retry_filename,
        file_bytes=file_bytes,
        size_bytes=len(file_bytes),
        retried_from_job_id=job_id,
    )


@router.get("/import-status/{job_id}", response_model=ImportJobStatusResponse)
def get_import_status(
    job_id: str,
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    repo = ImportRepository(db)
    job = repo.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")
    if job.created_by_user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Import job not found")

    percentage = 0.0
    if job.total_rows > 0:
        percentage = round((job.processed_rows / job.total_rows) * 100, 2)

    errors = [
        ImportErrorItem(row=item.row_number, error=item.error_message)
        for item in repo.fetch_errors(job_id, limit=5000)
    ]

    failed_report_download_url = None
    if job.failed_report_path:
        failed_report_download_url = f"/api/admin/import-errors/{job_id}/download"

    return ImportJobStatusResponse(
        job_id=job.id,
        state=job.status,
        total_rows=job.total_rows,
        processed_rows=job.processed_rows,
        success_count=job.success_count,
        failed_count=job.failed_count,
        percentage_completed=percentage,
        estimated_time_remaining_seconds=job.eta_seconds,
        errors=errors,
        failed_report_download_url=failed_report_download_url,
    )


@router.get("/import-errors/{job_id}/download")
def download_import_errors(
    job_id: str,
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    repo = ImportRepository(db)
    job = repo.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")
    if job.created_by_user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Import job not found")

    if not job.failed_report_path:
        raise HTTPException(status_code=404, detail="No failed row report available for this job")

    if not os.path.exists(job.failed_report_path):
        raise HTTPException(status_code=404, detail="Failed row report file no longer exists")

    return FileResponse(
        path=job.failed_report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"import_{job.id}_failed_rows.xlsx",
    )
