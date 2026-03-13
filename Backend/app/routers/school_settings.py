import csv
import io
import json
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.security import get_current_user_with_roles, has_any_role
from app.database import get_db
from app.models.associations import program_department_association
from app.models.department import Department
from app.models.program import Program
from app.models.role import Role
from app.models.school import School, SchoolAuditLog, SchoolSetting
from app.models.user import SSGProfile, StudentProfile, User as UserModel, UserRole
from app.schemas.school_settings import (
    SchoolAuditLogResponse,
    SchoolSettingsResponse,
    SchoolSettingsUpdate,
    UserImportRowResult,
    UserImportSummary,
)
from app.schemas.user import RoleEnum, SSGPositionEnum
from app.services.email_service import EmailDeliveryError, send_welcome_email
from app.utils.passwords import generate_secure_password

router = APIRouter(prefix="/school-settings", tags=["school-settings"])

EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
VALID_ROLE_VALUES = {role.value for role in RoleEnum}
SSG_POSITION_MAP = {pos.value.lower(): pos.value for pos in SSGPositionEnum}
ROW_HEADERS = [
    "email",
    "first_name",
    "middle_name",
    "last_name",
    "roles",
    "student_id",
    "department_id",
    "program_id",
    "year_level",
    "ssg_position",
]
ROLE_ALIASES = {
    "school_it": "school_IT",
    "school it": "school_IT",
    "event_organizer": "event-organizer",
    "event organizer": "event-organizer",
}


def _ensure_school_it_or_admin(current_user: UserModel) -> None:
    if not has_any_role(current_user, ["school_IT", "admin"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="School IT or admin privileges required",
        )


def _resolve_current_school(db: Session, current_user: UserModel) -> School:
    user_school_id = getattr(current_user, "school_id", None)
    if user_school_id is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User is not assigned to a school.",
        )

    school = db.query(School).filter(School.id == user_school_id).first()
    if school is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assigned school was not found.",
        )

    return school


def _get_or_create_school_settings(db: Session, school_id: int) -> SchoolSetting:
    settings = db.query(SchoolSetting).filter(SchoolSetting.school_id == school_id).first()
    if settings:
        return settings

    settings = SchoolSetting(school_id=school_id)
    db.add(settings)
    db.commit()
    db.refresh(settings)
    return settings


def _build_settings_response(school: School, settings: SchoolSetting) -> SchoolSettingsResponse:
    return SchoolSettingsResponse(
        school_id=school.id,
        school_name=school.school_name or school.name,
        logo_url=school.logo_url,
        primary_color=school.primary_color or settings.primary_color,
        secondary_color=school.secondary_color or settings.secondary_color,
        accent_color=settings.accent_color,
    )


def _write_audit_log(
    db: Session,
    school_id: int,
    actor_user_id: Optional[int],
    action: str,
    status_value: str,
    details: Optional[dict] = None,
) -> None:
    log_entry = SchoolAuditLog(
        school_id=school_id,
        actor_user_id=actor_user_id,
        action=action,
        status=status_value,
        details=json.dumps(details or {}, default=str),
    )
    db.add(log_entry)


def _normalize_role_name(role_name: str) -> str:
    normalized = role_name.strip().lower()
    return ROLE_ALIASES.get(normalized, normalized)


def _normalize_headers(raw_headers: List[str]) -> List[str]:
    normalized = []
    for header in raw_headers:
        clean = str(header or "").strip().lower().replace("-", "_").replace(" ", "_")
        normalized.append(clean)
    return normalized


def _parse_roles(raw_roles: str) -> Tuple[List[str], List[str]]:
    errors: List[str] = []
    role_values: List[str] = []

    if not raw_roles.strip():
        return role_values, ["roles is required"]

    parts = [part.strip() for part in re.split(r"[;,|]", raw_roles) if part.strip()]
    if not parts:
        return role_values, ["roles is required"]

    for part in parts:
        normalized = _normalize_role_name(part)
        if normalized not in VALID_ROLE_VALUES:
            errors.append(f"invalid role '{part}'")
            continue
        role_values.append(normalized)

    deduped_roles = list(dict.fromkeys(role_values))
    return deduped_roles, errors


def _load_import_rows(upload_file: UploadFile) -> List[dict]:
    filename = upload_file.filename or "uploaded_file"
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    raw_content = upload_file.file.read()
    if not raw_content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    rows: List[dict] = []

    if extension == "csv":
        try:
            decoded = raw_content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail=f"Failed to decode CSV file: {exc}") from exc

        reader = csv.DictReader(io.StringIO(decoded))
        if reader.fieldnames is None:
            raise HTTPException(status_code=400, detail="CSV file is missing headers")
        normalized_headers = _normalize_headers(list(reader.fieldnames))
        for row_index, raw_row in enumerate(reader, start=2):
            parsed_row = {
                normalized_headers[i]: str(value or "").strip()
                for i, (_, value) in enumerate(raw_row.items())
            }
            parsed_row["_row_number"] = row_index
            rows.append(parsed_row)
        return rows

    if extension in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="openpyxl is required to process Excel files",
            ) from exc

        workbook = load_workbook(io.BytesIO(raw_content), data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        normalized_headers = _normalize_headers(headers)
        for row_index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            parsed_row = {}
            for idx, value in enumerate(row_values):
                key = normalized_headers[idx] if idx < len(normalized_headers) else f"col_{idx}"
                parsed_row[key] = str(value or "").strip()
            parsed_row["_row_number"] = row_index
            rows.append(parsed_row)
        return rows

    raise HTTPException(
        status_code=400,
        detail="Unsupported file format. Please upload a .xlsx or .csv file.",
    )


def _to_optional_int(raw_value: str, field_name: str, errors: List[str]) -> Optional[int]:
    value = (raw_value or "").strip()
    if not value:
        return None
    try:
        return int(value)
    except ValueError:
        errors.append(f"{field_name} must be a valid integer")
        return None


@router.get("/me", response_model=SchoolSettingsResponse)
def get_my_school_settings(
    current_user: UserModel = Depends(get_current_user_with_roles),
    db: Session = Depends(get_db),
):
    _ensure_school_it_or_admin(current_user)
    school = _resolve_current_school(db, current_user)
    settings = _get_or_create_school_settings(db, school.id)
    return _build_settings_response(school, settings)


@router.put("/me", response_model=SchoolSettingsResponse)
def update_my_school_settings(
    payload: SchoolSettingsUpdate,
    current_user: UserModel = Depends(get_current_user_with_roles),
    db: Session = Depends(get_db),
):
    _ensure_school_it_or_admin(current_user)
    school = _resolve_current_school(db, current_user)
    settings = _get_or_create_school_settings(db, school.id)

    changes = {}

    if payload.school_name is not None:
        new_school_name = payload.school_name.strip()
        current_name = school.school_name or school.name
        if new_school_name != current_name:
            changes["school_name"] = {"from": current_name, "to": new_school_name}
        school.name = new_school_name
        school.school_name = new_school_name

    if payload.logo_url is not None:
        new_logo_url = payload.logo_url.strip() or None
        if new_logo_url != school.logo_url:
            changes["logo_url"] = {"from": school.logo_url, "to": new_logo_url}
        school.logo_url = new_logo_url

    if payload.primary_color is not None:
        current_primary = school.primary_color or settings.primary_color
        if payload.primary_color != current_primary:
            changes["primary_color"] = {"from": current_primary, "to": payload.primary_color}
        school.primary_color = payload.primary_color
        settings.primary_color = payload.primary_color
    if payload.secondary_color is not None:
        current_secondary = school.secondary_color or settings.secondary_color
        if payload.secondary_color != current_secondary:
            changes["secondary_color"] = {"from": current_secondary, "to": payload.secondary_color}
        school.secondary_color = payload.secondary_color
        settings.secondary_color = payload.secondary_color
    if payload.accent_color is not None:
        if payload.accent_color != settings.accent_color:
            changes["accent_color"] = {"from": settings.accent_color, "to": payload.accent_color}
        settings.accent_color = payload.accent_color

    settings.updated_by_user_id = current_user.id

    _write_audit_log(
        db=db,
        school_id=school.id,
        actor_user_id=current_user.id,
        action="branding_update",
        status_value="success",
        details={"changes": changes, "changed": bool(changes)},
    )

    db.commit()
    db.refresh(school)
    db.refresh(settings)

    return _build_settings_response(school, settings)


@router.get("/me/audit-logs", response_model=List[SchoolAuditLogResponse])
def list_school_audit_logs(
    limit: int = 50,
    current_user: UserModel = Depends(get_current_user_with_roles),
    db: Session = Depends(get_db),
):
    _ensure_school_it_or_admin(current_user)
    school = _resolve_current_school(db, current_user)
    logs = (
        db.query(SchoolAuditLog)
        .filter(SchoolAuditLog.school_id == school.id)
        .order_by(SchoolAuditLog.created_at.desc())
        .limit(max(1, min(limit, 200)))
        .all()
    )
    return logs


@router.get("/me/users/import-template")
def download_user_import_template(
    current_user: UserModel = Depends(get_current_user_with_roles),
):
    _ensure_school_it_or_admin(current_user)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is required to generate the Excel template",
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Users"

    sheet.append(ROW_HEADERS)
    sheet.append(
        [
            "student1@example.com",
            "Jane",
            "",
            "Doe",
            "student",
            "STUD-0001",
            "1",
            "1",
            "1",
            "",
        ]
    )
    sheet.append(
        [
            "ssg1@example.com",
            "John",
            "",
            "Smith",
            "ssg,event-organizer",
            "",
            "",
            "",
            "",
            "President",
        ]
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    reference = workbook.create_sheet("Reference")
    reference.append(["Field", "Notes"])
    reference.append(["roles", "Comma-separated values: student, ssg, event-organizer"])
    reference.append(["student_id", "Required when roles include student"])
    reference.append(["department_id/program_id/year_level", "Required when roles include student"])
    reference.append(["ssg_position", "Required when roles include ssg"])
    reference.append(["temporary_password", "Auto-generated by the system and emailed to user"])
    reference.append(["valid_ssg_positions", ", ".join(pos.value for pos in SSGPositionEnum)])
    for cell in reference[1]:
        cell.font = Font(bold=True)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="user_import_template.xlsx"'},
    )


@router.post("/me/users/import", response_model=UserImportSummary)
def import_users_from_excel(
    file: UploadFile = File(...),
    current_user: UserModel = Depends(get_current_user_with_roles),
    db: Session = Depends(get_db),
):
    _ensure_school_it_or_admin(current_user)
    school = _resolve_current_school(db, current_user)

    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    rows = _load_import_rows(file)
    non_empty_rows = [
        row for row in rows if any((str(value or "").strip() for key, value in row.items() if key != "_row_number"))
    ]

    roles = db.query(Role).all()
    role_map = {role.name: role for role in roles}
    missing_role_values = [name for name in VALID_ROLE_VALUES if name not in role_map]
    if missing_role_values:
        _write_audit_log(
            db=db,
            school_id=school.id,
            actor_user_id=current_user.id,
            action="user_import_attempt",
            status_value="failed",
            details={
                "filename": file.filename,
                "error": f"Missing role definitions: {', '.join(sorted(missing_role_values))}",
            },
        )
        db.commit()
        raise HTTPException(
            status_code=500,
            detail=f"Missing role definitions in database: {', '.join(sorted(missing_role_values))}",
        )

    global_email_owner = {
        email.lower(): owner_school_id
        for email, owner_school_id in db.query(UserModel.email, UserModel.school_id).all()
    }
    existing_emails = {
        email for email, owner_school_id in global_email_owner.items() if owner_school_id == school.id
    }
    global_student_owner = {
        student_id: owner_school_id
        for student_id, owner_school_id in (
            db.query(StudentProfile.student_id, UserModel.school_id)
            .join(UserModel, StudentProfile.user_id == UserModel.id)
            .all()
        )
        if student_id
    }
    existing_student_ids = {
        student_id
        for student_id, owner_school_id in global_student_owner.items()
        if owner_school_id == school.id
    }

    department_ids = {
        department_id for (department_id,) in db.query(Department.id).all()
    }
    program_ids = {
        program_id for (program_id,) in db.query(Program.id).all()
    }
    program_department_map: Dict[int, set] = {}
    for program_id, department_id in db.execute(
        select(
            program_department_association.c.program_id,
            program_department_association.c.department_id,
        )
    ).all():
        program_department_map.setdefault(program_id, set()).add(department_id)

    seen_emails = set()
    seen_student_ids = set()
    results: List[UserImportRowResult] = []
    created_count = 0
    email_failed_count = 0

    for row in non_empty_rows:
        row_number = int(row.get("_row_number", 0))
        errors: List[str] = []

        email = str(row.get("email", "")).strip().lower()
        first_name = str(row.get("first_name", "")).strip()
        middle_name = str(row.get("middle_name", "")).strip() or None
        last_name = str(row.get("last_name", "")).strip()
        raw_roles = str(row.get("roles", "")).strip()
        student_id = str(row.get("student_id", "")).strip() or None
        ssg_position = str(row.get("ssg_position", "")).strip()

        if not email:
            errors.append("email is required")
        elif not EMAIL_REGEX.match(email):
            errors.append("email is invalid")
        elif email in seen_emails:
            errors.append("duplicate email in uploaded file")
        elif email in existing_emails:
            errors.append("email already exists in this school")
        elif email in global_email_owner and global_email_owner[email] != school.id:
            errors.append("email already exists in another school")

        if not first_name:
            errors.append("first_name is required")
        if not last_name:
            errors.append("last_name is required")

        parsed_roles, role_errors = _parse_roles(raw_roles)
        errors.extend(role_errors)
        disallowed_school_scoped_roles = {"admin", "school_IT"}
        forbidden_roles = sorted(role for role in parsed_roles if role in disallowed_school_scoped_roles)
        if forbidden_roles:
            errors.append(
                "school-scoped import cannot assign roles: " + ", ".join(forbidden_roles)
            )

        department_id = _to_optional_int(str(row.get("department_id", "")), "department_id", errors)
        program_id = _to_optional_int(str(row.get("program_id", "")), "program_id", errors)
        year_level = _to_optional_int(str(row.get("year_level", "")), "year_level", errors)

        if "student" in parsed_roles:
            if not student_id:
                errors.append("student_id is required for student role")
            elif student_id in seen_student_ids:
                errors.append("duplicate student_id in uploaded file")
            elif student_id in existing_student_ids:
                errors.append("student_id already exists in this school")
            elif student_id in global_student_owner and global_student_owner[student_id] != school.id:
                errors.append("student_id already exists in another school")

            if department_id is None:
                errors.append("department_id is required for student role")
            elif department_id not in department_ids:
                errors.append(f"department_id '{department_id}' does not exist")

            if program_id is None:
                errors.append("program_id is required for student role")
            elif program_id not in program_ids:
                errors.append(f"program_id '{program_id}' does not exist")

            if year_level is None:
                errors.append("year_level is required for student role")
            elif year_level < 1 or year_level > 5:
                errors.append("year_level must be between 1 and 5")

            if (
                department_id is not None
                and program_id is not None
                and program_id in program_department_map
                and department_id not in program_department_map[program_id]
            ):
                errors.append("program_id is not linked to the provided department_id")

        if "ssg" in parsed_roles:
            if not ssg_position:
                errors.append("ssg_position is required for ssg role")
            elif ssg_position.lower() not in SSG_POSITION_MAP:
                errors.append("ssg_position is invalid")

        if errors:
            results.append(
                UserImportRowResult(
                    row_number=row_number,
                    email=email or None,
                    status="failed",
                    errors=errors,
                    user_id=None,
                )
            )
            continue

        try:
            temporary_password = generate_secure_password(min_length=10, max_length=14)
            db_user = UserModel(
                email=email,
                school_id=school.id,
                first_name=first_name,
                middle_name=middle_name,
                last_name=last_name,
                must_change_password=True,
            )
            db_user.set_password(temporary_password)
            db.add(db_user)
            db.flush()

            for role_name in parsed_roles:
                db.add(UserRole(user_id=db_user.id, role_id=role_map[role_name].id))

            if "student" in parsed_roles:
                db.add(
                    StudentProfile(
                        user_id=db_user.id,
                        school_id=school.id,
                        student_id=student_id,
                        department_id=department_id,
                        program_id=program_id,
                        year_level=year_level,
                    )
                )

            if "ssg" in parsed_roles:
                db.add(
                    SSGProfile(
                        user_id=db_user.id,
                        position=SSG_POSITION_MAP[ssg_position.lower()],
                    )
                )

            db.commit()
            created_count += 1
            seen_emails.add(email)
            existing_emails.add(email)
            global_email_owner[email] = school.id
            if student_id:
                seen_student_ids.add(student_id)
                existing_student_ids.add(student_id)
                global_student_owner[student_id] = school.id

            row_result_errors: List[str] = []
            try:
                send_welcome_email(
                    recipient_email=email,
                    temporary_password=temporary_password,
                    first_name=first_name,
                    system_name=school.school_name or school.name,
                )
            except EmailDeliveryError as email_exc:
                email_failed_count += 1
                row_result_errors.append(f"welcome email not sent: {email_exc}")

            results.append(
                UserImportRowResult(
                    row_number=row_number,
                    email=email,
                    status="created",
                    errors=row_result_errors,
                    user_id=db_user.id,
                )
            )
        except IntegrityError:
            db.rollback()
            results.append(
                UserImportRowResult(
                    row_number=row_number,
                    email=email,
                    status="failed",
                    errors=["database uniqueness check failed for this row"],
                    user_id=None,
                )
            )
        except Exception as exc:
            db.rollback()
            results.append(
                UserImportRowResult(
                    row_number=row_number,
                    email=email,
                    status="failed",
                    errors=[f"unexpected error: {exc}"],
                    user_id=None,
                )
            )

    failed_count = len(results) - created_count
    audit_status = "success" if failed_count == 0 else "partial_success"
    if created_count == 0 and failed_count > 0:
        audit_status = "failed"

    _write_audit_log(
        db=db,
        school_id=school.id,
        actor_user_id=current_user.id,
        action="user_import_attempt",
        status_value=audit_status,
        details={
            "filename": file.filename,
            "total_rows": len(non_empty_rows),
            "created_count": created_count,
            "failed_count": failed_count,
            "email_sent_count": max(created_count - email_failed_count, 0),
            "email_failed_count": email_failed_count,
        },
    )
    db.commit()

    return UserImportSummary(
        filename=file.filename,
        total_rows=len(non_empty_rows),
        created_count=created_count,
        failed_count=failed_count,
        results=results,
    )
