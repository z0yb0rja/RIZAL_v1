from __future__ import annotations

import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.database import SessionLocal
from app.models.school import SchoolAuditLog
from app.models.user import User
from app.repositories.import_repository import ImportRepository
from app.services.import_validation_service import (
    EXPECTED_HEADERS,
    HeaderValidationError,
    ValidationContext,
    sanitize_excel_output,
    validate_and_transform_row,
    validate_headers,
)
from app.utils.passwords import generate_secure_password, hash_password_bcrypt
from app.worker.celery_app import celery_app
from app.models.department import Department
from app.models.program import Program
from app.models.school import School


logger = logging.getLogger(__name__)


class StudentImportService:
    def __init__(self) -> None:
        self.settings = get_settings()

    def process_job(self, job_id: str) -> None:
        with SessionLocal() as db:
            repo = ImportRepository(db)
            job = repo.get_job(job_id)
            if not job:
                logger.error("Import job not found", extra={"job_id": job_id})
                return

            repo.mark_processing(job_id)
            db.commit()

        failed_report_path: str | None = None

        lock_db: Session | None = None
        try:
            lock_db = SessionLocal()
            lock_repo = ImportRepository(lock_db)
            lock_repo.lock_import_processing()
            lock_db.commit()

            failed_report_path = self._process_streaming(job_id)

            with SessionLocal() as db:
                repo = ImportRepository(db)
                repo.mark_completed(job_id, failed_report_path=failed_report_path)
                job = repo.get_job(job_id)
                if job:
                    self._append_import_audit_log(
                        db,
                        job=job,
                        status_value="success",
                        details={
                            "job_id": job.id,
                            "total_rows": job.total_rows,
                            "processed_rows": job.processed_rows,
                            "success_count": job.success_count,
                            "failed_count": job.failed_count,
                            "failed_report_path": failed_report_path,
                        },
                    )
                db.commit()

        except Exception as exc:
            logger.exception("Import processing failed", extra={"job_id": job_id})
            with SessionLocal() as db:
                repo = ImportRepository(db)
                safe_error = self._safe_error_message(exc)
                repo.mark_failed(job_id, safe_error)
                job = repo.get_job(job_id)
                if job:
                    self._append_import_audit_log(
                        db,
                        job=job,
                        status_value="failed",
                        details={
                            "job_id": job.id,
                            "error": safe_error,
                        },
                    )
                db.commit()
        finally:
            if lock_db is not None:
                try:
                    lock_repo = ImportRepository(lock_db)
                    lock_repo.unlock_import_processing()
                    lock_db.commit()
                finally:
                    lock_db.close()

    def _process_streaming(self, job_id: str) -> str | None:
        settings = self.settings
        start_time = datetime.utcnow()

        with SessionLocal() as db:
            repo = ImportRepository(db)
            job = repo.get_job(job_id)
            if not job:
                raise RuntimeError("Import job not found")
            file_path = job.stored_file_path
            target_school_id = job.target_school_id

        if not os.path.exists(file_path):
            raise FileNotFoundError("Uploaded file was not found on server")

        validation_context = self._build_validation_context(target_school_id)

        failed_report_dir = Path(settings.import_storage_dir) / "reports"
        failed_report_dir.mkdir(parents=True, exist_ok=True)

        failed_workbook = Workbook(write_only=True)
        failed_sheet = failed_workbook.create_sheet("Failed Rows")
        failed_sheet.append(EXPECTED_HEADERS + ["Error"])

        row_buffer: List[dict] = []
        error_buffer: List[dict] = []

        processed_rows = 0
        success_count = 0
        failed_count = 0
        total_rows = 0

        with SessionLocal() as db:
            repo = ImportRepository(db)
            student_role_id = repo.get_student_role_id()
            db.commit()

        workbook = None
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            sheet = workbook.active

            estimated_total = max((sheet.max_row or 1) - 1, 0)
            total_rows = estimated_total

            row_iter = sheet.iter_rows(values_only=True)
            header_row = next(row_iter, None)
            if header_row is None:
                raise HeaderValidationError("Uploaded Excel file is empty")
            validate_headers(header_row)

            for row_number, row_values in enumerate(row_iter, start=2):
                transformed, row_errors, raw_row_data = validate_and_transform_row(
                    row_number=row_number,
                    row_values=row_values,
                    context=validation_context,
                )
                processed_rows += 1

                if row_errors:
                    failed_count += 1
                    self._append_failed_row(failed_sheet, raw_row_data, "; ".join(row_errors))
                    error_buffer.append(
                        {
                            "row": row_number,
                            "error": "; ".join(row_errors),
                            "row_data": raw_row_data,
                        }
                    )
                else:
                    temporary_password = generate_secure_password()
                    transformed["temporary_password"] = temporary_password
                    transformed["password_hash"] = hash_password_bcrypt(temporary_password)
                    transformed["raw_row_data"] = raw_row_data
                    row_buffer.append(transformed)

                if len(row_buffer) >= settings.import_chunk_size:
                    batch_success_count, batch_failed_count, batch_errors = self._flush_batch(
                        job_id=job_id,
                        row_buffer=row_buffer,
                        student_role_id=student_role_id,
                    )
                    success_count += batch_success_count
                    failed_count += batch_failed_count
                    for error_item in batch_errors:
                        self._append_failed_row(
                            failed_sheet,
                            error_item.get("row_data", {}),
                            error_item.get("error", "unknown error"),
                        )
                    error_buffer.extend(batch_errors)
                    row_buffer = []

                if len(error_buffer) >= 1000:
                    with SessionLocal() as db:
                        repo = ImportRepository(db)
                        repo.add_errors(job_id, error_buffer)
                        db.commit()
                    error_buffer = []

                if processed_rows % 500 == 0:
                    self._update_progress(
                        job_id=job_id,
                        total_rows=total_rows,
                        processed_rows=processed_rows,
                        success_count=success_count,
                        failed_count=failed_count,
                        start_time=start_time,
                    )

            if row_buffer:
                batch_success_count, batch_failed_count, batch_errors = self._flush_batch(
                    job_id=job_id,
                    row_buffer=row_buffer,
                    student_role_id=student_role_id,
                )
                success_count += batch_success_count
                failed_count += batch_failed_count
                for error_item in batch_errors:
                    self._append_failed_row(
                        failed_sheet,
                        error_item.get("row_data", {}),
                        error_item.get("error", "unknown error"),
                    )
                error_buffer.extend(batch_errors)

            if error_buffer:
                with SessionLocal() as db:
                    repo = ImportRepository(db)
                    repo.add_errors(job_id, error_buffer)
                    db.commit()

            self._update_progress(
                job_id=job_id,
                total_rows=processed_rows if total_rows == 0 else total_rows,
                processed_rows=processed_rows,
                success_count=success_count,
                failed_count=failed_count,
                start_time=start_time,
            )

            if failed_count > 0:
                report_path = failed_report_dir / f"{job_id}_failed_rows.xlsx"
                failed_workbook.save(str(report_path))
                return str(report_path)

            return None

        except Exception:
            raise
        finally:
            if workbook is not None:
                workbook.close()

    def _safe_error_message(self, exc: Exception) -> str:
        if isinstance(exc, (HeaderValidationError, FileNotFoundError, InvalidFileException, BadZipFile)):
            return str(exc)
        return "Import processing failed. Please validate the file and try again."

    def _append_import_audit_log(
        self,
        db: Session,
        *,
        job,
        status_value: str,
        details: dict,
    ) -> None:
        school_id = (
            db.query(User.school_id)
            .filter(User.id == job.created_by_user_id)
            .scalar()
        )
        if school_id is None:
            return

        db.add(
            SchoolAuditLog(
                school_id=school_id,
                actor_user_id=job.created_by_user_id,
                action="student_bulk_import_result",
                status=status_value,
                details=json.dumps(details, default=str),
            )
        )

    def _flush_batch(
        self,
        *,
        job_id: str,
        row_buffer: List[dict],
        student_role_id: int,
    ) -> tuple[int, int, List[dict]]:
        with SessionLocal() as db:
            repo = ImportRepository(db)
            success_rows, batch_errors = repo.bulk_insert_students(row_buffer, student_role_id)
            db.commit()

        for row in success_rows:
            celery_app.send_task(
                "app.worker.tasks.send_student_welcome_email",
                args=[
                    job_id,
                    row["user_id"],
                    row["email"],
                    row["temporary_password"],
                    row.get("first_name"),
                ],
            )

        return len(success_rows), len(batch_errors), batch_errors

    def _build_validation_context(self, target_school_id: int) -> ValidationContext:
        with SessionLocal() as db:
            school_exists = (
                db.query(School.id)
                .filter(School.id == target_school_id)
                .first()
            )
            if not school_exists:
                raise RuntimeError("Target school does not exist")
            department_lookup = {
                department_name.strip().lower(): department_id
                for department_id, department_name in db.query(Department.id, Department.name).all()
            }
            course_lookup = {
                course_name.strip().lower(): course_id
                for course_id, course_name in db.query(Program.id, Program.name).all()
            }

        return ValidationContext(
            target_school_id=target_school_id,
            department_lookup=department_lookup,
            course_lookup=course_lookup,
        )

    def _append_failed_row(self, failed_sheet, row_data: dict, error_message: str) -> None:
        row_values = [
            sanitize_excel_output(str(row_data.get(header, ""))) for header in EXPECTED_HEADERS
        ]
        failed_sheet.append(row_values + [sanitize_excel_output(error_message)])

    def _update_progress(
        self,
        *,
        job_id: str,
        total_rows: int,
        processed_rows: int,
        success_count: int,
        failed_count: int,
        start_time: datetime,
    ) -> None:
        elapsed_seconds = max((datetime.utcnow() - start_time).total_seconds(), 1)
        processing_rate = processed_rows / elapsed_seconds

        eta_seconds = None
        if total_rows > processed_rows and processing_rate > 0:
            eta_seconds = int((total_rows - processed_rows) / processing_rate)

        with SessionLocal() as db:
            repo = ImportRepository(db)
            repo.update_progress(
                job_id,
                total_rows=total_rows,
                processed_rows=processed_rows,
                success_count=success_count,
                failed_count=failed_count,
                eta_seconds=eta_seconds,
            )
            db.commit()
